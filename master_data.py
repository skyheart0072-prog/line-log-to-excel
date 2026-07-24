"""
master_data.py
โหลดข้อมูลอ้างอิง (master data) จาก data/master_data.xlsx
ใช้สำหรับให้ AI จับคู่ชื่อสาขา / ประเภทงาน / ทีม IT Support ที่ดูแล
"""

import os
import logging
from openpyxl import load_workbook

logger = logging.getLogger("line-log-bot")

MASTER_DATA_PATH = os.getenv("MASTER_DATA_PATH", "data/master_data.xlsx")


def load_master_data() -> dict:
    """
    คืนค่าเป็น dict:
    {
        "branches": [{"code": "CJ0344", "name": "ชุมชนหนองปรือ", "project": "ATM"}, ...],
        "issue_teams": [{"keyword": "เหรียญติด", "it_support": "Flook"}, ...]
    }
    ถ้าไฟล์ไม่มี จะคืน dict ว่างและ log เตือน (ระบบยังทำงานได้ แต่ AI จะจับคู่ไม่ได้แม่นเท่าที่ควร)
    """
    if not os.path.exists(MASTER_DATA_PATH):
        logger.warning(f"ไม่พบไฟล์ master data ที่ {MASTER_DATA_PATH} - ใช้ค่าว่างไปก่อน")
        return {"branches": [], "issue_teams": []}

    wb = load_workbook(MASTER_DATA_PATH, data_only=True)

    branches = []
    if "Master_Branch" in wb.sheetnames:
        ws = wb["Master_Branch"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, name, project = (row + (None, None, None))[:3]
            if code:
                branches.append({"code": code, "name": name, "project": project})

    issue_teams = []
    if "Master_Issue" in wb.sheetnames:
        ws = wb["Master_Issue"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            keyword, it_support = (row + (None, None))[:2]
            if keyword:
                issue_teams.append({"keyword": keyword, "it_support": it_support})

    return {"branches": branches, "issue_teams": issue_teams}


def is_user_allowed(user_id: str, allowed_ids: set) -> bool:
    """ตรวจสอบสิทธิ์ผู้ใช้งานจาก LINE User ID"""
    if not allowed_ids:
        # ถ้ายังไม่ได้ตั้งค่า ALLOWED_USER_IDS เลย ให้ผ่านทุกคนไปก่อน (โหมด dev เท่านั้น)
        logger.warning("ALLOWED_USER_IDS ยังไม่ถูกตั้งค่า - อนุญาตทุกคนชั่วคราว")
        return True
    return user_id in allowed_ids
